import os
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Paths base de plantillas
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANTILLA_PATH = os.path.join(BASE_DIR, "Inventarios", "PLANTILLAS", "plantilla_base_formato_cliente.xlsm")
PLANTILLA_PATH_MASTER = os.path.join(BASE_DIR, "Inventarios", "PLANTILLAS", "plantilla_base.xlsm")

# Configuración UI principal
st.set_page_config(page_title="Consolidador de Inventarios", page_icon="🧾", layout="centered")
st.title("Consolidador de Inventarios por Marca")

# Instrucciones
col1, col2 = st.columns([3, 1])
with col1:
    st.markdown("### 🧠 ¿Cómo usar esta herramienta?")
    st.markdown("""
    1. **Usa la plantilla oficial** para recolectar los datos.
    2. Asegúrate de que todos los archivos subidos contengan la columna **Marca** (pueden ser distintas).
    3. Sube los archivos y la herramienta detectará todas las marcas.
    4. Selecciona las marcas que quieres consolidar.
    5. Para cada marca seleccionada, ingresa los datos fijos del cliente y selecciona mes/año.
    6. Se mostrará siempre 1 registro por fila, aunque algunas Cajas sean 0 o estén vacías.
    7. Las fechas se guardan en el consolidado en formato `YYYY-MM-DD` (sin hora).
    """)
with col2:
    with open(PLANTILLA_PATH, "rb") as f:
        st.download_button("📥 Plantilla oficial", f.read(), "plantilla_base.xlsm", mime="application/vnd.ms-excel.sheet.macroEnabled.12")
    with open(PLANTILLA_PATH, "rb") as f:
        st.download_button("📥 Plantilla clientes", f.read(), "plantilla_base_formato_cliente.xlsm", mime="application/vnd.ms-excel.sheet.macroEnabled.12")

# Inicialización de estado
if "file_uploader_key" not in st.session_state:
    st.session_state["file_uploader_key"] = "initial"

# Botón de reinicio
if st.button("🔁 Limpiar todo y comenzar de nuevo"):
    st.session_state.clear()
    st.session_state["file_uploader_key"] = str(datetime.now())
    st.rerun()

# Carga de archivos
uploaded_files = st.file_uploader(
    "📂 Sube uno o más archivos Excel con la plantilla oficial",
    type=["xlsm", "xlsx"],
    accept_multiple_files=True,
    key=st.session_state["file_uploader_key"]
)

registros_finales = []
marcas_detectadas = set()
archivos_por_marca = {}  
df_por_archivo = {}      

# Procesa cada archivo cargado
if uploaded_files:
    st.success(f"Has subido {len(uploaded_files)} archivo(s).")
    for file in uploaded_files:
        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip()

            columnas_base = ["Nombre Comercial", "Tipo de cliente", "Marca", "Codigo de producto", "Descripción"]
            if not all(col in df.columns for col in columnas_base):
                st.error(f"❌ {file.name} no tiene las columnas necesarias.")
                continue

            # Detecta marcas en el archivo
            for marca in df["Marca"].dropna().unique():
                archivos_por_marca.setdefault(marca, []).append(file.name)
                marcas_detectadas.add(marca)

            df_por_archivo[file.name] = df
            cajas_cols = [c for c in df.columns if 'Cajas' in c]
            fechas_cols = [c for c in df.columns if 'Fecha' in c]

            # Estructura base por registro
            for _, row in df.iterrows():
                base_data = {col: row[col] for col in columnas_base}
                base_data["__archivo_origen__"] = file.name
                for i in range(min(len(cajas_cols), len(fechas_cols))):
                    base_data[f"Cajas_{i+1}"] = row[cajas_cols[i]]
                    base_data[f"Fecha_{i+1}"] = row[fechas_cols[i]]
                registros_finales.append(base_data)

        except Exception as e:
            st.error(f"❌ Error procesando {file.name}: {e}")

# Selección de marcas a consolidar
marcas_seleccionadas = []
if marcas_detectadas:
    st.markdown("### ✅ Selecciona las marcas a consolidar")
    for marca in sorted(marcas_detectadas):
        if st.checkbox(marca, value=True, key=f"chk_{marca}"):
            marcas_seleccionadas.append(marca)

# Formularios de configuración
datos_por_marca = {}
datos_por_archivo = {}

if marcas_seleccionadas:
    st.markdown("### 📝 Datos de consolidación")
    for marca in marcas_seleccionadas:
        with st.expander(f"📌 Configuración para {marca}", expanded=True):
            if len(archivos_por_marca.get(marca, [])) > 1:
                modo = st.radio(
                    f"⚙️ Valores para {marca}",
                    ["Mismos valores para todos", "Valores distintos por archivo"],
                    key=f"modo_{marca}"
                )
            else:
                modo = "Mismos valores para todos"

            # Configuración unificada
            if modo == "Mismos valores para todos":
                col1, col2 = st.columns(2)
                with col1:
                    mes = st.selectbox(f"📆 Mes ({marca})", [
                        "Enero","Febrero","Marzo","Abril","Mayo","Junio",
                        "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"
                    ], key=f"mes_{marca}")
                    ruta = st.text_input(f"🛣️ Ruta ({marca})", key=f"ruta_{marca}")
                    codigo_cliente = st.text_input(f"🔢 Código de cliente ({marca})", key=f"codigo_cliente_{marca}")
                with col2:
                    año = st.number_input(f"📅 Año ({marca})", min_value=2023, max_value=2100,
                                           value=datetime.now().year, key=f"año_{marca}")
                    zona = st.text_input(f"📍 Zona ({marca})", key=f"zona_{marca}")
                    tienda = st.text_input(f"🏪 Nombre de la tienda ({marca})", key=f"tienda_{marca}")

                datos_por_marca[marca] = {
                    "modo": "mismo",
                    "Mes": mes, "Ruta": ruta, "Codigo de cliente": codigo_cliente,
                    "Año": año, "Zona": zona, "Nombre de la tienda": tienda
                }
            else:
                # Configuración individual por archivo
                datos_por_marca[marca] = {"modo": "distinto"}
                for archivo in archivos_por_marca[marca]:
                    st.markdown(f"**Archivo:** {archivo}")
                    col1, col2 = st.columns(2)
                    with col1:
                        mes = st.selectbox(f"📆 Mes ({archivo})", [
                            "Enero","Febrero","Marzo","Abril","Mayo","Junio",
                            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"
                        ], key=f"mes_{marca}_{archivo}")
                        ruta = st.text_input(f"🛣️ Ruta ({archivo})", key=f"ruta_{marca}_{archivo}")
                        codigo_cliente = st.text_input(f"🔢 Código de cliente ({archivo})", key=f"codigo_cliente_{marca}_{archivo}")
                    with col2:
                        año = st.number_input(f"📅 Año ({archivo})", min_value=2023, max_value=2100,
                                               value=datetime.now().year, key=f"año_{marca}_{archivo}")
                        zona = st.text_input(f"📍 Zona ({archivo})", key=f"zona_{marca}_{archivo}")
                        tienda = st.text_input(f"🏪 Nombre de la tienda ({archivo})", key=f"tienda_{marca}_{archivo}")

                    datos_por_archivo[archivo] = {
                        "Mes": mes, "Ruta": ruta, "Codigo de cliente": codigo_cliente,
                        "Año": año, "Zona": zona, "Nombre de la tienda": tienda
                    }

# Generación de archivos consolidados
if registros_finales and datos_por_marca and st.button("📦 Generar archivo consolidado por marca"):
    with st.spinner("⏳ Generando..."):
        df_consolidado = pd.DataFrame(registros_finales)
        st.session_state.downloads_por_marca = []

        for marca, config in datos_por_marca.items():
            df_marca = df_consolidado[df_consolidado["Marca"] == marca].copy()
            if df_marca.empty:
                continue

            # Formatea fechas
            for col in df_marca.columns:
                if col.startswith("Fecha_"):
                    df_marca[col] = pd.to_datetime(df_marca[col], errors="coerce").dt.strftime("%Y-%m-%d")

            # Asigna datos fijos o por archivo
            if config["modo"] == "mismo":
                for col, valor in config.items():
                    if col != "modo":
                        df_marca[col] = valor
            else:
                for idx, row in df_marca.iterrows():
                    for col, valor in datos_por_archivo[row["__archivo_origen__"]].items():
                        df_marca.at[idx, col] = valor

            # Orden de columnas final
            orden_columnas = [
                "Mes","Ruta","Zona","Codigo de cliente","Nombre de la tienda",
                "Nombre Comercial","Tipo de cliente","Marca","Codigo de producto","Descripción",
                "Cajas_1","Fecha_1","Cajas_2","Fecha_2","Cajas_3","Fecha_3","Cajas_4","Fecha_4"
            ]
            df_marca = df_marca[[c for c in orden_columnas if c in df_marca.columns]]

            # Inserta datos en plantilla
            wb = load_workbook(PLANTILLA_PATH_MASTER, keep_vba=True)
            ws = wb.active
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            for row in dataframe_to_rows(df_marca, index=False, header=False):
                ws.append(row)

            # Prepara descarga
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            file_name = f"{config.get('Ruta', 'MULTI_Ruta')}_Inventario_{marca}_{config.get('Mes', 'MULTI_Mes')}_{config.get('Año', 'MULTI_Año')}.xlsm"
            st.session_state.downloads_por_marca.append((file_name, output))

    st.success("✅ Archivos generados.")

# Botones de descarga final
if "downloads_por_marca" in st.session_state and st.session_state.downloads_por_marca:
    st.markdown("### 📦 Archivos disponibles:")
    for file_name, file_data in st.session_state.downloads_por_marca:
        st.download_button(
            label=f"📥 Descargar {file_name}",
            data=file_data,
            file_name=file_name,
            mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            key=file_name
        )
